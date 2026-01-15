from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple
import xml.etree.ElementTree as ET


# =========================
# Util
# =========================
def _br_money(v: float) -> str:
    s = f"{v:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def _br_num(v: float, casas: int = 2) -> str:
    s = f"{v:,.{casas}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _to_float(s: str | None) -> float:
    if not s:
        return 0.0
    s = str(s).strip()
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _strip_namespaces(root: ET.Element) -> ET.Element:
    for el in root.iter():
        if "}" in el.tag:
            el.tag = el.tag.split("}", 1)[1]
    return root


def _zip_iter_files(zip_bytes: bytes) -> List[Tuple[str, bytes]]:
    out: List[Tuple[str, bytes]] = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as z:
        for name in z.namelist():
            if name.endswith("/"):
                continue
            out.append((name, z.read(name)))
    return out


def _findtext(root: ET.Element, *paths: str, default: str = "") -> str:
    """
    Tenta vários caminhos e devolve o primeiro texto encontrado.
    """
    for p in paths:
        el = root.find(p)
        if el is not None and (el.text or "").strip():
            return (el.text or "").strip()
    return default


def _fmt_data(dh: str) -> str:
    """
    Converte ISO para dd/mm/aaaa quando possível.
    Ex: 2026-01-15T10:20:30-03:00 -> 15/01/2026
    """
    if not dh:
        return ""
    dh = dh.strip()
    if len(dh) >= 10 and dh[4] == "-" and dh[7] == "-":
        yyyy = dh[0:4]
        mm = dh[5:7]
        dd = dh[8:10]
        return f"{dd}/{mm}/{yyyy}"
    return dh


# =========================
# cClass (Excel)
# =========================
def carregar_cclass_lista(xlsx_path: str = "data/Tabela-cClass.xlsx") -> List[Dict[str, str]]:
    candidatos = [
        Path(xlsx_path),
        Path("Tabela-cClass.xlsx"),
        Path("data") / "Tabela-cClass.xlsx",
    ]
    caminho = None
    for c in candidatos:
        if c.exists():
            caminho = c
            break
    if caminho is None:
        return []

    import openpyxl

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active

    header = []
    for col in range(1, ws.max_column + 1):
        header.append(str(ws.cell(1, col).value or "").strip().lower())

    def find_col(*names: str) -> int | None:
        for n in names:
            n = n.lower()
            for i, h in enumerate(header, start=1):
                if h == n:
                    return i
        return None

    col_code = find_col("cclass", "codigo", "código", "code", "grupo/código") or 1
    col_desc = find_col("descricao", "descrição", "desc") or 2

    lista: List[Dict[str, str]] = []
    for row in range(2, ws.max_row + 1):
        code = str(ws.cell(row, col_code).value or "").strip()
        desc = str(ws.cell(row, col_desc).value or "").strip()
        if code:
            lista.append({"code": code, "desc": desc})
    return lista


# =========================
# Mapa de descrição do cClass (compat)
# =========================
#
# O seu app (app.py) às vezes importa `cclass_desc_map` (lowercase) e/ou
# `cClass_desc_map` (camel). Python diferencia maiúsculas/minúsculas.
# Para não travar o deploy no Render, expomos os dois nomes.

cClass_desc_map: Dict[str, str] = {}
# Alias compatível (não é cópia; aponta para o mesmo dict)
cclass_desc_map = cClass_desc_map


def get_cclass_desc_map(xlsx_path: str = "data/Tabela-cClass.xlsx") -> Dict[str, str]:
    """Retorna o mapa {cClass: descrição} carregado do Excel (quando disponível)."""
    global cClass_desc_map, cclass_desc_map
    if cClass_desc_map:
        return cClass_desc_map

    try:
        lista = carregar_cclass_lista(xlsx_path)
        cClass_desc_map = {row["code"].strip(): row["desc"].strip() for row in lista if row.get("code")}
    except Exception:
        # Em ambientes sem arquivo/permissão, mantém vazio para não quebrar.
        cClass_desc_map = {}

    # Mantém o alias sincronizado
    cclass_desc_map = cClass_desc_map
    return cClass_desc_map


# =========================
# Regras texto (cClass;CFOP por linha)
# =========================
def parse_regras_texto(txt: str | None) -> Dict[str, str]:
    """
    Entrada:
      0600101;5102
      110201;5102
    """
    regras: Dict[str, str] = {}
    if not txt:
        return regras
    for line in txt.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if ";" in line:
            a, b = line.split(";", 1)
            a = a.strip()
            b = b.strip()
            if a and b:
                regras[a] = b
        elif "," in line:
            a, b = line.split(",", 1)
            a = a.strip()
            b = b.strip()
            if a and b:
                regras[a] = b
    return regras


# =========================
# Processamento Lote ZIP -> ZIP (mantido)
# =========================
def _aplicar_regras_xml_str(xml_str: str, regras: Dict[str, str], remover_desc: bool, remover_outros: bool) -> str:
    """
    Aplica CFOP conforme cClass (quando encontra <cClass>XXXX</cClass> no item),
    e remove algumas tags se marcado.
    """
    # CFOP por cClass: tenta inserir/atualizar tag CFOP no mesmo "bloco" do item
    # (simples e funcional com regex para o seu caso)
    def repl(match):
        cclass = match.group(1)
        cfop = regras.get(cclass)
        bloco = match.group(0)
        if not cfop:
            return bloco

        if re.search(r"<CFOP>.*?</CFOP>", bloco, flags=re.DOTALL):
            bloco = re.sub(r"<CFOP>.*?</CFOP>", f"<CFOP>{cfop}</CFOP>", bloco, flags=re.DOTALL)
            return bloco

        # se não tem CFOP, insere após cClass
        bloco = bloco.replace(f"<cClass>{cclass}</cClass>", f"<cClass>{cclass}</cClass><CFOP>{cfop}</CFOP>")
        return bloco

    xml_str = re.sub(r"<cClass>(\d+)</cClass>.*?(?=</det>|</Item>|</item>|</prod>|</Produto>|</produto>)",
                     repl, xml_str, flags=re.DOTALL)

    # Remover tags (exemplos)
    if remover_desc:
        xml_str = re.sub(r"<vDesc>.*?</vDesc>", "", xml_str, flags=re.DOTALL)
    if remover_outros:
        xml_str = re.sub(r"<vOutro>.*?</vOutro>", "", xml_str, flags=re.DOTALL)

    return xml_str


def processar_lote_zip(zip_bytes: bytes, regras: Dict[str, str], remover_desc: bool, remover_outros: bool) -> bytes:
    mem_out = io.BytesIO()
    with zipfile.ZipFile(mem_out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in _zip_iter_files(zip_bytes):
            if not name.lower().endswith(".xml"):
                zout.writestr(name, data)
                continue
            try:
                s = data.decode("utf-8", errors="ignore")
                s2 = _aplicar_regras_xml_str(s, regras, remover_desc, remover_outros)
                zout.writestr(name, s2.encode("utf-8"))
            except Exception:
                zout.writestr(name, data)
    return mem_out.getvalue()


# =========================
# Nota Única (mantido) - parse_nfcom (para resultado.html)
# =========================
def parse_nfcom(xml_bytes: bytes) -> Dict:
    root = ET.fromstring(xml_bytes)
    root = _strip_namespaces(root)

    ide = root.find(".//ide")
    emit = root.find(".//emit")
    dest = root.find(".//dest")
    total = root.find(".//total")

    nNF = _findtext(ide, ".//nNF", default="")
    serie = _findtext(ide, ".//serie", default="")
    dhEmi = _findtext(ide, ".//dhEmi", ".//dEmi", default="")
    emit_nome = _findtext(emit, ".//xNome", default="")
    dest_nome = _findtext(dest, ".//xNome", default="")

    itens = []
    for det in root.findall(".//det"):
        cClass = _findtext(det, ".//cClass", default="")
        xProd = _findtext(det, ".//xProd", default="")
        vProd = _to_float(_findtext(det, ".//vProd", default="0"))
        itens.append({
            "cClass": cClass,
            "xProd": xProd,
            "vProd": vProd,
            "vProd_br": _br_money(vProd),
        })

    vNF = _to_float(_findtext(total, ".//vNF", default="0"))
    return {
        "nNF": nNF,
        "serie": serie,
        "dhEmi": dhEmi,
        "dhEmi_fmt": _fmt_data(dhEmi),
        "emit_nome": emit_nome,
        "dest_nome": dest_nome,
        "vNF": vNF,
        "vNF_br": _br_money(vNF),
        "itens": itens,
    }


# =========================
# CSV (mantido)
# =========================
def gerar_csv_de_zip(zip_bytes: bytes, mapping: List[Tuple[str, str]]) -> bytes:
    # mapping: [(header, tag), ...]
    bio = io.StringIO()
    w = csv.writer(bio, delimiter=";")
    w.writerow([m[0] for m in mapping])

    for name, data in _zip_iter_files(zip_bytes):
        if not name.lower().endswith(".xml"):
            continue
        root = ET.fromstring(data)
        root = _strip_namespaces(root)
        row = []
        for _, campo in mapping:
            el = root.find(f".//{campo}")
            row.append(el.text.strip() if el is not None and el.text else "")
        w.writerow(row)

    return bio.getvalue().encode("utf-8-sig")


# =========================
# RESUMO (NOVO/ATUALIZADO): Itens + notas por item
# =========================
@dataclass
class ItemResumo:
    cclass: str
    xprod: str
    vprod: float
    nnf: str
    cnf: str
    dest_nome: str
    dhemi: str


def parse_nfcom_itens(xml_bytes: bytes) -> List[ItemResumo]:
    """
    Extrai itens (xProd, cClass, vProd) + dados da nota (nNF, emit xNome, dhEmi).
    """
    root = ET.fromstring(xml_bytes)
    root = _strip_namespaces(root)

    ide = root.find(".//ide")
    dest = root.find(".//dest")

    nnf = _findtext(ide, ".//nNF", default="")
    cnf = _findtext(ide, ".//cNF", default="")
    dhemi = _findtext(ide, ".//dhEmi", ".//dEmi", default="")
    dest_nome = _findtext(dest, ".//xNome", default="")

    itens: List[ItemResumo] = []
    for det in root.findall(".//det"):
        cclass = _findtext(det, ".//cClass", default="").strip()
        xprod = _findtext(det, ".//xProd", default="").strip()
        vprod = _to_float(_findtext(det, ".//vProd", default="0"))

        # ignora linhas vazias
        if not cclass and not xprod and vprod == 0:
            continue

        itens.append(ItemResumo(
            cclass=cclass,
            xprod=xprod,
            vprod=vprod,
            nnf=nnf,
            cnf=cnf,
            dest_nome=dest_nome,
            dhemi=dhemi
        ))
    return itens


def gerar_resumo_de_zip(zip_bytes: bytes, desc_map: Dict[str, str] | None = None) -> Dict:
    """
    Retorna um dict pronto pro template:
      - total_arquivos, total_geral_br
      - linhas (por cClass)
      - labels/valores (pro gráfico)
      - itens_linhas (top 50) com accordion:
          cada linha tem .notas = lista de notas daquele item
    """
    total_arquivos = 0
    total_geral = 0.0

    # por cClass
    por_cclass: Dict[str, Dict[str, float]] = {}

    # por item (xProd + cClass)
    por_item: Dict[Tuple[str, str], Dict[str, float]] = {}

    # notas por item:
    # key_item -> key_nota -> soma vprod daquele item naquela nota
    # key_nota = (nNF, cNF, dest/xNome, dhEmi)
    por_item_notas: Dict[Tuple[str, str], Dict[Tuple[str, str, str, str], float]] = {}

    for name, data in _zip_iter_files(zip_bytes):
        if not name.lower().endswith(".xml"):
            continue

        try:
            itens = parse_nfcom_itens(data)
        except Exception:
            continue

        total_arquivos += 1

        for it in itens:
            v = float(it.vprod)
            total_geral += v

            # cClass
            por_cclass.setdefault(it.cclass, {"qtd_itens": 0, "v_total": 0.0})
            por_cclass[it.cclass]["qtd_itens"] += 1
            por_cclass[it.cclass]["v_total"] += v

            # item agregado
            key_item = (it.xprod or "(sem descrição)", it.cclass or "")
            por_item.setdefault(key_item, {"qtd_itens": 0, "v_total": 0.0})
            por_item[key_item]["qtd_itens"] += 1
            por_item[key_item]["v_total"] += v

            # notas por item
            key_nota = (it.nnf or "", it.cnf or "", it.dest_nome or "", it.dhemi or "")
            por_item_notas.setdefault(key_item, {})
            por_item_notas[key_item][key_nota] = por_item_notas[key_item].get(key_nota, 0.0) + v

    # monta linhas por cClass (ordenadas por valor)
    linhas = []
    for cclass, agg in por_cclass.items():
        v_total = float(agg["v_total"])
        qtd = int(agg["qtd_itens"])
        pct = (v_total / total_geral * 100.0) if total_geral > 0 else 0.0
        linhas.append({
            "cClass": cclass,
            "qtd_itens": qtd,
            "v_total": v_total,
            "v_total_br": _br_money(v_total),
            "pct": pct,
            "pct_br": f"{pct:.2f}".replace(".", ","),
        })
    linhas.sort(key=lambda x: x["v_total"], reverse=True)

    # dados do gráfico (top 12)
    top = linhas[:12]
    labels = [x["cClass"] for x in top]
    valores = [x["v_total"] for x in top]

    # monta itens_linhas (top 50) com notas (accordion)
    itens_linhas = []
    for (xprod, cclass), agg in por_item.items():
        v_total = float(agg["v_total"])
        qtd = int(agg["qtd_itens"])
        pct = (v_total / total_geral * 100.0) if total_geral > 0 else 0.0

        notas_map = por_item_notas.get((xprod, cclass), {})
        notas_list = []
        for (nnf, cnf, dest_nome, dhemi), vnota in notas_map.items():
            notas_list.append({
                "nNF": nnf,
                "cNF": cnf,
                "xNome": dest_nome,
                "dhEmi": dhemi,
                "dhEmi_fmt": _fmt_data(dhemi),
                "vProd": vnota,
                "vProd_br": _br_money(float(vnota)),
            })
        # opcional: ordenar notas por valor do item na nota (desc)
        notas_list.sort(key=lambda x: x["vProd"], reverse=True)

        itens_linhas.append({
            "item": xprod,
            "cClass": cclass,
            "qtd_itens": qtd,
            "v_total": v_total,
            "v_total_br": _br_money(v_total),
            "pct": pct,
            "pct_br": f"{pct:.2f}".replace(".", ","),
            "notas": notas_list,
        })

    itens_linhas.sort(key=lambda x: x["v_total"], reverse=True)

    return {
        "total_arquivos": total_arquivos,
        "total_geral": total_geral,
        "total_geral_br": _br_money(total_geral),
        "linhas": linhas,
        "labels": labels,
        "valores": valores,
        "itens_linhas": itens_linhas,
    }


# =========================
# Nota Única (VISUALIZAÇÃO) - compatível com resultado.html
# =========================
def gerar_dados_nota_xml(xml_bytes: bytes) -> Dict:
    """Extrai campos para o template resultado.html.

    A NFCom varia bastante por provedor/leiaute, então aqui a regra é:
    - procurar as tags mais comuns (com vários caminhos)
    - não quebrar se alguma parte não existir
    """

    root = ET.fromstring(xml_bytes)
    root = _strip_namespaces(root)

    ide = root.find(".//ide")
    emit = root.find(".//emit")
    dest = root.find(".//dest")
    total = root.find(".//total")

    def join_endereco(ender: ET.Element | None) -> Tuple[str, str]:
        if ender is None:
            return "", ""
        lgr = _findtext(ender, ".//xLgr", default="")
        nro = _findtext(ender, ".//nro", default="")
        bairro = _findtext(ender, ".//xBairro", default="")
        mun = _findtext(ender, ".//xMun", default="")
        uf = _findtext(ender, ".//UF", default="")
        cep = _findtext(ender, ".//CEP", default="")
        linha1 = " ".join([x for x in [lgr, ("nº " + nro) if nro else ""] if x]).strip()
        linha2 = " - ".join([x for x in [bairro, mun, uf] if x]).strip()
        if cep:
            linha2 = (linha2 + (" " if linha2 else "") + "CEP " + cep).strip()
        return linha1, linha2

    # Cabeçalho
    nNF = _findtext(ide, ".//nNF", default="")
    serie = _findtext(ide, ".//serie", default="")
    cNF = _findtext(ide, ".//cNF", default="")
    dhEmi = _findtext(ide, ".//dhEmi", ".//dEmi", default="")

    emit_nome = _findtext(emit, ".//xNome", default="")
    emit_fantasia = _findtext(emit, ".//xFant", default="")
    emit_cnpj = _findtext(emit, ".//CNPJ", ".//CPF", default="")
    emit_ie = _findtext(emit, ".//IE", default="")
    emit_ender = emit.find(".//enderEmit") if emit is not None else None
    emit_l1, emit_l2 = join_endereco(emit_ender)

    dest_nome = _findtext(dest, ".//xNome", default="")
    dest_doc = _findtext(dest, ".//CNPJ", ".//CPF", default="")
    dest_ender = dest.find(".//enderDest") if dest is not None else None
    dest_l1, dest_l2 = join_endereco(dest_ender)

    # Totais (tenta achar em vários lugares)
    vNF = _to_float(_findtext(total, ".//vNF", default="0"))
    bc_total = _to_float(_findtext(total, ".//vBC", ".//ICMS//vBC", default="0"))
    icms_total = _to_float(_findtext(total, ".//vICMS", ".//ICMS//vICMS", default="0"))
    pis_total = _to_float(_findtext(total, ".//vPIS", ".//PIS//vPIS", default="0"))
    cofins_total = _to_float(_findtext(total, ".//vCOFINS", ".//COFINS//vCOFINS", default="0"))
    fust_total = _to_float(_findtext(total, ".//vFUST", default="0"))
    funttel_total = _to_float(_findtext(total, ".//vFUNTTEL", default="0"))

    # Itens
    itens: List[Dict] = []
    for det in root.findall(".//det"):
        cClass = _findtext(det, ".//cClass", default="")
        xProd = _findtext(det, ".//xProd", default="")
        cfop = _findtext(det, ".//CFOP", ".//cfop", ".//cCFOP", default="")
        un = _findtext(det, ".//uCom", ".//uTrib", ".//uUnid", default="")
        qtd = _to_float(_findtext(det, ".//qCom", ".//qTrib", ".//qUnid", default="0"))
        vUnit = _to_float(_findtext(det, ".//vUnCom", ".//vUnTrib", ".//vUnid", default="0"))
        vTotal = _to_float(_findtext(det, ".//vProd", default="0"))

        v_pis = _to_float(_findtext(det, ".//PIS//vPIS", ".//vPIS", default="0"))
        v_cofins = _to_float(_findtext(det, ".//COFINS//vCOFINS", ".//vCOFINS", default="0"))
        pis_cofins = v_pis + v_cofins

        bc_icms = _to_float(_findtext(det, ".//ICMS//vBC", ".//vBC", default="0"))
        aliq_icms = _to_float(_findtext(det, ".//ICMS//pICMS", ".//pICMS", default="0"))
        v_icms = _to_float(_findtext(det, ".//ICMS//vICMS", ".//vICMS", default="0"))

        itens.append({
            "cClass": cClass,
            "cfop": cfop,
            "xProd": xProd,
            "un": un,
            "qtd": qtd,
            "qtd_fmt": _br_num(qtd, 2) if qtd else "",
            "vUnit": vUnit,
            "vUnit_fmt": _br_money(vUnit) if vUnit else "",
            "vTotal": vTotal,
            "vTotal_fmt": _br_money(vTotal) if vTotal else "",
            "pis_cofins": pis_cofins,
            "pis_cofins_fmt": _br_money(pis_cofins) if pis_cofins else "",
            "bc_icms": bc_icms,
            "bc_icms_fmt": _br_money(bc_icms) if bc_icms else "",
            "aliq_icms": aliq_icms,
            "aliq_icms_fmt": _br_num(aliq_icms, 2) if aliq_icms else "",
            "v_icms": v_icms,
            "v_icms_fmt": _br_money(v_icms) if v_icms else "",
        })

    return {
        "nNF": nNF,
        "serie": serie,
        "dhEmi": dhEmi,
        "dhEmi_fmt": _fmt_data(dhEmi),
        "emit_nome": emit_nome,
        "emit_fantasia": emit_fantasia,
        "emit_cnpj": emit_cnpj,
        "emit_ie": emit_ie,
        "emit_endereco_linha1": emit_l1,
        "emit_endereco_linha2": emit_l2,
        "dest_nome": dest_nome,
        "dest_doc": dest_doc,
        "dest_endereco_linha1": dest_l1,
        "dest_endereco_linha2": dest_l2,
        # você pediu contrato = cNF
        "contrato": cNF,
        "total": vNF,
        "total_fmt": _br_money(vNF),
        "total_pagar_fmt": _br_money(vNF),
        "bc_total_fmt": _br_money(bc_total) if bc_total else "",
        "icms_total_fmt": _br_money(icms_total) if icms_total else "",
        "pis_total_fmt": _br_money(pis_total) if pis_total else "",
        "cofins_total_fmt": _br_money(cofins_total) if cofins_total else "",
        "fust_total_fmt": _br_money(fust_total) if fust_total else "",
        "funttel_total_fmt": _br_money(funttel_total) if funttel_total else "",
        "itens": itens,
    }
